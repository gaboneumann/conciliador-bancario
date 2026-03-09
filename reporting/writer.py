"""
writer.py — Escritura de archivos Excel de resultado (v2.2).

Responsabilidad: tomar el DataFrame clasificado y producir
los archivos Excel de salida con formato aplicado.

Estructura visual de conciliacion_resultado.xlsx:
    Pestaña "Conciliación"      → todas las transacciones con match y resultado
    Pestaña "Resumen"           → totales, porcentajes y diferencia de saldo
    Pestaña "Hallazgos_Criticos"→ ranking por RUT de partidas sin conciliar (PRD v2.2)

CAMBIOS v2.2:
─────────────────────────────────────────────────────────────────────────────
- _construir_hallazgos()        : genera DataFrame agrupado por RUT desde Manuales
- _escribir_hallazgos_criticos(): escribe la hoja con colores y alertas
- escribir_resultado()          : agrega la tercera pestaña al workbook
─────────────────────────────────────────────────────────────────────────────
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config.config import ARCHIVO_RESULTADO, ARCHIVO_SIN_CONCILIAR, OUTPUT_DIR
from reporting.formatter import (
    estilo_encabezado,
    estilo_encabezado_bloque,
    estilo_fila,
    estilo_hallazgo,
    estilo_numero,
    estilo_fecha,
    ANCHOS_RESULTADO,
    ANCHOS_SIN_CONCILIAR,
    ANCHOS_HALLAZGOS,
    BLOQUES_RESULTADO,
    BLOQUES_SIN_CONCILIAR,
    BLOQUES_HALLAZGOS,
)
from utils.logger import get_logger
from utils.exceptions import ConciliadorError

logger = get_logger(__name__)

# ─── Umbral de concentración (PRD v2.2) ──────────────────────────────────────
UMBRAL_CONCENTRACION = 0.20   # 20% — RUT que concentra más del 20% del error total


# ─── Verificación de archivo disponible ──────────────────────────────────────

def _verificar_archivo_disponible(ruta) -> None:
    if ruta.exists():
        try:
            ruta.unlink()
        except PermissionError:
            raise ConciliadorError(
                f"El archivo '{ruta.name}' está abierto en Excel. "
                f"Ciérralo e intenta de nuevo."
            )


# ─── Definición de columnas v2 ────────────────────────────────────────────────

ENCABEZADOS_RESULTADO = [
    "Fecha Operación", "Fecha Valor", "Glosa Cartola",
    "RUT Cartola", "Monto Cartola", "Nº Documento", "Banco",
    "Fecha Contable", "Glosa Libro",
    "RUT Libro", "Monto Libro", "Nº Referencia", "Nº Comprobante", "Código Tx",
    "Tipo Match", "Certeza", "Regla Aplicada",
    "Dif. Monto", "Dif. Días",
    "Flag Conciliación", "Flag IVA",
    "Días Antigüedad", "Tramo Antigüedad", "Acción Recomendada",
]

COLUMNAS_RESULTADO = [
    "fecha_operacion_cartola", "fecha_valor_cartola", "glosa_cartola",
    "rut_cartola", "monto_cartola", "nro_documento_cartola", "banco_cartola",
    "fecha_contable_libro", "glosa_libro",
    "rut_libro", "monto_libro", "nro_referencia_libro",
    "nro_comprobante_libro", "codigo_tx_libro",
    "tipo_match", "certeza", "regla_aplicada",
    "diff_monto", "diff_dias",
    "flag_conciliacion", "flag_iva",
    "dias_antiguedad", "tramo_antiguedad", "accion_recomendada",
]

ENCABEZADOS_SIN_CONCILIAR = [
    "Fecha Operación", "Fecha Valor", "Glosa Cartola",
    "RUT Cartola", "Monto Cartola", "Nº Documento", "Banco",
    "Motivo", "Monto Más Cercano", "Dif. Monto Cercano",
]

COLUMNAS_SIN_CONCILIAR = [
    "fecha_operacion_cartola", "fecha_valor_cartola", "glosa_cartola",
    "rut_cartola", "monto_cartola", "nro_documento_cartola", "banco_cartola",
    "motivo", "monto_cercano", "diff_monto_cercano",
]

ENCABEZADOS_HALLAZGOS = [
    "RUT",
    "Glosa Frecuente",
    "Cantidad Partidas",
    "Monto Total Pendiente",
    "Motivo Principal",
    "Antigüedad Máxima (días)",
    "% sobre Total Error",
    "Alerta",
]

COLS_MONTO = {
    "monto_cartola", "monto_libro", "diff_monto",
    "monto_cercano", "diff_monto_cercano",
}
COLS_FECHA = {
    "fecha_operacion_cartola", "fecha_valor_cartola",
    "fecha_contable_libro", "fecha_cercana",
}


# ─── Función interna: fila de encabezados agrupados ──────────────────────────

def _escribir_fila_bloques(ws, bloques: list) -> None:
    ws.row_dimensions[1].height = 22

    for bloque in bloques:
        col_ini = bloque["col_inicio"]
        col_fin = bloque["col_fin"]
        nombre  = bloque["nombre"]
        tipo    = bloque["bloque"]

        if col_ini < col_fin:
            ws.merge_cells(
                start_row=1, start_column=col_ini,
                end_row=1,   end_column=col_fin,
            )

        celda         = ws.cell(row=1, column=col_ini, value=nombre)
        estilo        = estilo_encabezado_bloque(tipo)
        celda.font    = estilo["font"]
        celda.fill    = estilo["fill"]
        celda.alignment = estilo["alignment"]
        celda.border  = estilo["border"]


# ─── Función interna: escritura de hoja estándar ─────────────────────────────

def _escribir_hoja(ws, df, columnas, encabezados, anchos, bloques) -> None:
    ws.sheet_view.showGridLines = False

    _escribir_fila_bloques(ws, bloques)

    ws.row_dimensions[2].height = 36
    estilo_enc = estilo_encabezado()

    for col_idx, nombre in enumerate(encabezados, start=1):
        celda           = ws.cell(row=2, column=col_idx, value=nombre)
        celda.font      = estilo_enc["font"]
        celda.fill      = estilo_enc["fill"]
        celda.alignment = estilo_enc["alignment"]
        celda.border    = estilo_enc["border"]

    for row_idx, (_, fila) in enumerate(df[columnas].iterrows(), start=3):
        tipo_match        = fila.get("tipo_match", "Manual")
        flag_conciliacion = fila.get("flag_conciliacion", "")
        estilo            = estilo_fila(tipo_match, flag_conciliacion=flag_conciliacion)

        for col_idx, col_nombre in enumerate(columnas, start=1):
            valor = fila[col_nombre]
            if pd.isna(valor):
                valor = None

            celda           = ws.cell(row=row_idx, column=col_idx, value=valor)
            celda.font      = estilo["font"]
            celda.fill      = estilo["fill"]
            celda.alignment = estilo["alignment"]
            celda.border    = estilo["border"]

            if col_nombre in COLS_MONTO and valor is not None:
                celda.number_format = estilo_numero()["number_format"]
                celda.alignment     = estilo_numero()["alignment"]
            elif col_nombre in COLS_FECHA and valor is not None:
                celda.number_format = estilo_fecha()["number_format"]
                celda.alignment     = estilo_fecha()["alignment"]

    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = ancho

    ws.freeze_panes = "A3"


# ─── Función interna: construir DataFrame de hallazgos ───────────────────────

def _construir_hallazgos(df_resultado: pd.DataFrame) -> pd.DataFrame:
    """
    Agrupa las partidas Manuales por RUT y construye el ranking de hallazgos.

    Columnas producidas:
        rut, glosa_frecuente, cantidad_partidas, monto_total,
        motivo_principal, antiguedad_max, pct_error, alerta, tramo_max

    Returns:
        DataFrame ordenado por antiguedad_max DESC (Críticos primero).
    """
    manuales = df_resultado[df_resultado["tipo_match"] == "Manual"].copy()

    if manuales.empty:
        return pd.DataFrame()

    monto_total_error = manuales["monto_cartola"].abs().sum()

    filas = []
    for rut, grupo in manuales.groupby("rut_cartola"):

        # Glosa más frecuente del RUT
        glosa_frecuente = (
            grupo["glosa_cartola"]
            .dropna()
            .mode()
            .iloc[0] if not grupo["glosa_cartola"].dropna().empty else ""
        )

        # Motivo predominante
        if "motivo" in grupo.columns:
            motivo_counts  = grupo["motivo"].value_counts()
            motivo_principal = motivo_counts.index[0] if not motivo_counts.empty else "Sin diagnóstico"
            # Traducir a etiqueta corta para el reporte
            mapa_motivo = {
                "Fecha coincide pero monto no encontrado":  "Omisión",
                "Monto coincide pero fecha fuera de rango": "Corte",
                "Posible Neto vs Bruto (×1.19)":            "IVA",
                "Transacción ausente en libro auxiliar":    "Ausente",
            }
            motivo_principal = mapa_motivo.get(motivo_principal, motivo_principal)
        else:
            motivo_principal = "Sin diagnóstico"

        cantidad       = len(grupo)
        monto_rut      = grupo["monto_cartola"].sum()
        antiguedad_max = grupo["dias_antiguedad"].max()
        tramo_max      = grupo.loc[grupo["dias_antiguedad"].idxmax(), "tramo_antiguedad"]
        pct_error      = (
            abs(monto_rut) / monto_total_error * 100
            if monto_total_error != 0 else 0
        )
        alerta = "⚠️ Concentración >20%" if pct_error > UMBRAL_CONCENTRACION * 100 else ""

        filas.append({
            "rut":               rut,
            "glosa_frecuente":   glosa_frecuente,
            "cantidad_partidas": cantidad,
            "monto_total":       round(monto_rut, 0),
            "motivo_principal":  motivo_principal,
            "antiguedad_max":    antiguedad_max,
            "pct_error":         round(pct_error, 1),
            "alerta":            alerta,
            "tramo_max":         tramo_max,
        })

    df_hallazgos = pd.DataFrame(filas)

    # Ordenar: Críticos primero, luego por antigüedad DESC
    orden_tramo = {"Crítico": 0, "En Observación": 1, "Vigente": 2}
    df_hallazgos["_orden_tramo"] = df_hallazgos["tramo_max"].map(orden_tramo).fillna(3)
    df_hallazgos = (
        df_hallazgos
        .sort_values(["_orden_tramo", "antiguedad_max"], ascending=[True, False])
        .drop(columns=["_orden_tramo"])
        .reset_index(drop=True)
    )

    return df_hallazgos


# ─── Función interna: escribir hoja Hallazgos_Criticos ───────────────────────

def _escribir_hallazgos_criticos(ws, df_hallazgos: pd.DataFrame) -> None:
    """
    Escribe la hoja Hallazgos_Criticos con ranking por RUT,
    colores de alerta y formato de números.

    Estructura:
        Fila 1 → bloque único "Hallazgos Críticos — Ranking por RUT"
        Fila 2 → encabezados de columna
        Fila 3+ → datos ordenados Críticos primero
    """
    ws.sheet_view.showGridLines = False

    # — Fila 1: bloque —
    _escribir_fila_bloques(ws, BLOQUES_HALLAZGOS)

    # — Fila 2: encabezados —
    ws.row_dimensions[2].height = 36
    estilo_enc = estilo_encabezado()
    for col_idx, nombre in enumerate(ENCABEZADOS_HALLAZGOS, start=1):
        celda           = ws.cell(row=2, column=col_idx, value=nombre)
        celda.font      = estilo_enc["font"]
        celda.fill      = estilo_enc["fill"]
        celda.alignment = estilo_enc["alignment"]
        celda.border    = estilo_enc["border"]

    if df_hallazgos.empty:
        ws.cell(row=3, column=1, value="Sin partidas manuales — conciliación completa ✅")
        return

    # — Fila 3+: datos —
    columnas_df = [
        "rut", "glosa_frecuente", "cantidad_partidas", "monto_total",
        "motivo_principal", "antiguedad_max", "pct_error", "alerta",
    ]

    for row_idx, (_, fila) in enumerate(df_hallazgos.iterrows(), start=3):
        alerta   = fila["alerta"]
        tramo    = fila["tramo_max"]
        estilo   = estilo_hallazgo(alerta, tramo)

        for col_idx, col_nombre in enumerate(columnas_df, start=1):
            valor = fila[col_nombre]
            if pd.isna(valor):
                valor = None

            celda           = ws.cell(row=row_idx, column=col_idx, value=valor)
            celda.font      = estilo["font"]
            celda.fill      = estilo["fill"]
            celda.alignment = estilo["alignment"]
            celda.border    = estilo["border"]

            # Formato monto
            if col_nombre == "monto_total" and valor is not None:
                celda.number_format = estilo_numero()["number_format"]
                celda.alignment     = estilo_numero()["alignment"]
            # Formato porcentaje
            elif col_nombre == "pct_error" and valor is not None:
                celda.number_format = '0.0"%"'
                celda.alignment     = estilo_numero()["alignment"]
            # Alinear números
            elif col_nombre in ("cantidad_partidas", "antiguedad_max") and valor is not None:
                celda.alignment = estilo_numero()["alignment"]

    # — Anchos —
    for letra, ancho in ANCHOS_HALLAZGOS.items():
        ws.column_dimensions[letra].width = ancho

    ws.freeze_panes = "A3"


# ─── Funciones públicas ───────────────────────────────────────────────────────

def escribir_resultado(df_resultado: pd.DataFrame, saldo: dict | None = None) -> None:
    """
    Escribe el reporte completo de conciliación con tres pestañas:
        1. Conciliación        → todas las transacciones
        2. Resumen             → totales y diferencia de saldo
        3. Hallazgos_Criticos  → ranking por RUT de partidas Manuales
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    _verificar_archivo_disponible(ARCHIVO_RESULTADO)

    wb = Workbook()

    # — Pestaña 1: Conciliación —
    ws_conc = wb.active
    ws_conc.title = "Conciliación"
    _escribir_hoja(
        ws_conc, df_resultado,
        COLUMNAS_RESULTADO, ENCABEZADOS_RESULTADO,
        ANCHOS_RESULTADO, BLOQUES_RESULTADO,
    )

    # — Pestaña 2: Resumen —
    ws_resumen = wb.create_sheet("Resumen")
    _escribir_resumen(ws_resumen, df_resultado, saldo=saldo)

    # — Pestaña 3: Hallazgos_Criticos —
    ws_hallazgos = wb.create_sheet("Hallazgos_Criticos")
    df_hallazgos = _construir_hallazgos(df_resultado)
    _escribir_hallazgos_criticos(ws_hallazgos, df_hallazgos)

    wb.save(ARCHIVO_RESULTADO)
    logger.info(f"Resultado guardado → {ARCHIVO_RESULTADO}")


def escribir_sin_conciliar(df_resultado: pd.DataFrame) -> None:
    """Escribe el reporte de partidas sin conciliar."""
    from conciliation.classifier import separar_sin_conciliar

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    _verificar_archivo_disponible(ARCHIVO_SIN_CONCILIAR)

    df_sin = separar_sin_conciliar(df_resultado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sin Conciliar"

    _escribir_hoja(
        ws, df_sin,
        COLUMNAS_SIN_CONCILIAR, ENCABEZADOS_SIN_CONCILIAR,
        ANCHOS_SIN_CONCILIAR, BLOQUES_SIN_CONCILIAR,
    )

    wb.save(ARCHIVO_SIN_CONCILIAR)
    logger.info(f"Sin conciliar guardado → {ARCHIVO_SIN_CONCILIAR} ({len(df_sin)} partidas)")


def _escribir_resumen(ws, df: pd.DataFrame, saldo: dict | None) -> None:
    """Escribe la hoja de resumen ejecutivo."""
    ws.sheet_view.showGridLines = False

    total     = len(df)
    exactos   = (df["tipo_match"] == "Exacto").sum()
    sugeridos = (df["tipo_match"] == "Sugerido").sum()
    manuales  = (df["tipo_match"] == "Manual").sum()
    pct_conc  = round((exactos + sugeridos) / total * 100, 1) if total > 0 else 0

    filas = [
        ("Resumen de Conciliación", ""),
        ("", ""),
        ("Total transacciones cartola", total),
        ("Match Exacto",               exactos),
        ("Match Sugerido",             sugeridos),
        ("Sin Match (Manual)",         manuales),
        ("% Conciliado",               f"{pct_conc}%"),
    ]

    if saldo:
        filas += [
            ("", ""),
            ("Diferencia de Saldo", ""),
            ("Saldo Cartola",  saldo["saldo_cartola"]),
            ("Saldo Libro",    saldo["saldo_libro"]),
            ("Diferencia",     saldo["diferencia"]),
            ("¿Cuadra?",       "Sí" if saldo["cuadra"] else "No"),
        ]

    estilo_enc  = estilo_encabezado()
    estilo_dato = estilo_fila("blanco")

    for row_idx, (etiqueta, valor) in enumerate(filas, start=1):
        celda_a = ws.cell(row=row_idx, column=1, value=etiqueta)
        celda_b = ws.cell(row=row_idx, column=2, value=valor)

        es_titulo = row_idx == 1 or etiqueta in ("Diferencia de Saldo",)

        if es_titulo and etiqueta:
            celda_a.font = estilo_enc["font"]
            celda_a.fill = estilo_enc["fill"]
            celda_b.fill = estilo_enc["fill"]
        else:
            celda_a.font = estilo_dato["font"]
            celda_b.font = estilo_dato["font"]

        if isinstance(valor, (int, float)) and etiqueta in (
            "Saldo Cartola", "Saldo Libro", "Diferencia"
        ):
            celda_b.number_format = estilo_numero()["number_format"]

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18