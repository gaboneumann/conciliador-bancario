"""
formatter.py — Estilos y colores para los Excel de salida (v2).

Responsabilidad única: definir y retornar objetos de estilo de openpyxl.
No escribe archivos ni manipula datos.

Colores de filas por tipo de match (v2):
    Exacto           → verde claro        #C6EFCE
    Sugerido         → amarillo claro     #FFEB9C
    Manual           → rojo claro         #FFC7CE
    flag_conciliacion→ azul claro         #BDD7EE  (prioridad sobre Sugerido)
    flag_iva         → verde agua         #E2EFDA

Colores de encabezados agrupados:
    Cartola Personal → azul oscuro        #1F4E79
    Libro del Banco  → verde oscuro       #1A5632
    Resultado        → gris oscuro        #404040
    Diagnóstico      → rojo oscuro        #7B2C2C
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─── Colores de filas ─────────────────────────────────────────────────────────

COLORES = {
    # v2
    "Exacto":        "C6EFCE",   # verde claro
    "Sugerido":      "FFEB9C",   # amarillo claro
    "Manual":        "FFC7CE",   # rojo claro
    "conciliacion":  "BDD7EE",   # azul claro  (flag Partida en Conciliación)
    "iva":           "E2EFDA",   # verde agua  (flag IVA)
    # v1 — mantenidos por retrocompatibilidad
    "exacto":        "C6EFCE",
    "parcial":       "FFEB9C",
    "sin_match":     "FFC7CE",
    # utilidades
    "encabezado":    "1F4E79",   # azul oscuro (encabezado simple)
    "blanco":        "FFFFFF",
    "gris_claro":    "F2F2F2",
}

# ─── Colores de bloques (encabezados agrupados) ───────────────────────────────

COLORES_BLOQUE = {
    "cartola":     "1F4E79",  # azul oscuro
    "libro":       "1A5632",  # verde oscuro
    "resultado":   "404040",  # gris oscuro
    "diagnostico": "7B2C2C",  # rojo oscuro
}

# ─── Fuente ───────────────────────────────────────────────────────────────────

FUENTE = "Arial"
TAMANO = 10


# ─── Constructores de estilo ──────────────────────────────────────────────────

def estilo_encabezado() -> dict:
    """Estilo para la fila de encabezados de columna (fila 2)."""
    return {
        "font": Font(
            name=FUENTE,
            size=TAMANO,
            bold=True,
            color="FFFFFF",
        ),
        "fill": PatternFill(
            fill_type="solid",
            start_color=COLORES["encabezado"],
        ),
        "alignment": Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        ),
        "border": _borde_fino(),
    }


def estilo_encabezado_bloque(bloque: str) -> dict:
    """
    Estilo para la fila superior de encabezados agrupados (fila 1).

    Args:
        bloque: "cartola", "libro", "resultado" o "diagnostico"

    Returns:
        Dict con font, fill, alignment y border
    """
    color = COLORES_BLOQUE.get(bloque, COLORES_BLOQUE["resultado"])

    return {
        "font": Font(
            name=FUENTE,
            size=TAMANO + 1,
            bold=True,
            color="FFFFFF",
        ),
        "fill": PatternFill(
            fill_type="solid",
            start_color=color,
        ),
        "alignment": Alignment(
            horizontal="center",
            vertical="center",
        ),
        "border": _borde_medio(),
    }


def estilo_fila(tipo_match: str, flag_conciliacion: str = "", flag_iva: str = "") -> dict:
    """
    Estilo para una fila de datos según su tipo de match y flags activos.

    Prioridad de color:
        1. flag_conciliacion activo → azul claro  (#BDD7EE)
        2. flag_iva activo          → verde agua  (#E2EFDA)
        3. tipo_match               → Exacto / Sugerido / Manual

    Args:
        tipo_match        : "Exacto", "Sugerido", "Manual" (v2) o valores v1
        flag_conciliacion : string no vacío activa el color azul
        flag_iva          : string no vacío activa el color verde agua

    Returns:
        Dict con font, fill, alignment y border listos para aplicar
    """
    if flag_conciliacion:
        color = COLORES["conciliacion"]
    elif flag_iva:
        color = COLORES["iva"]
    else:
        color = COLORES.get(tipo_match, COLORES["blanco"])

    return {
        "font": Font(name=FUENTE, size=TAMANO),
        "fill": PatternFill(fill_type="solid", start_color=color),
        "alignment": Alignment(vertical="center"),
        "border": _borde_fino(),
    }


def estilo_numero() -> dict:
    """Estilo adicional para celdas con montos."""
    return {
        "number_format": '#,##0_-;[Red](#,##0)',
        "alignment": Alignment(horizontal="right", vertical="center"),
    }


def estilo_fecha() -> dict:
    """Estilo adicional para celdas con fechas."""
    return {
        "number_format": "YYYY-MM-DD",
        "alignment": Alignment(horizontal="center", vertical="center"),
    }


# ─── Anchos de columna v2 ─────────────────────────────────────────────────────
# resultado    : 7 cartola + 7 libro + 10 resultado = 24 columnas
# sin_conciliar: 7 cartola + 3 diagnóstico          = 10 columnas

ANCHOS_RESULTADO = {
    # — Cartola (A–G) —
    "A": 14,   # fecha_operacion_cartola
    "B": 14,   # fecha_valor_cartola
    "C": 35,   # glosa_cartola
    "D": 18,   # rut_cartola
    "E": 16,   # monto_cartola
    "F": 18,   # nro_documento_cartola
    "G": 20,   # banco_cartola
    # — Libro (H–N) —
    "H": 14,   # fecha_contable_libro
    "I": 35,   # glosa_libro
    "J": 18,   # rut_libro
    "K": 16,   # monto_libro
    "L": 18,   # nro_referencia_libro
    "M": 16,   # nro_comprobante_libro
    "N": 14,   # codigo_tx_libro
    # — Resultado (O–X) —
    "O": 12,   # tipo_match
    "P": 12,   # certeza
    "Q": 28,   # regla_aplicada
    "R": 14,   # diff_monto
    "S": 12,   # diff_dias
    "T": 26,   # flag_conciliacion
    "U": 26,   # flag_iva
    "V": 14,   # dias_antiguedad
    "W": 18,   # tramo_antiguedad
    "X": 45,   # accion_recomendada
}

ANCHOS_SIN_CONCILIAR = {
    # — Cartola (A–G) —
    "A": 14,   # fecha_operacion_cartola
    "B": 14,   # fecha_valor_cartola
    "C": 35,   # glosa_cartola
    "D": 18,   # rut_cartola
    "E": 16,   # monto_cartola
    "F": 18,   # nro_documento_cartola
    "G": 20,   # banco_cartola
    # — Diagnóstico (H–J) —
    "H": 45,   # motivo
    "I": 16,   # monto_cercano
    "J": 16,   # diff_monto_cercano
}

# ─── Definición de bloques para encabezados agrupados (v2) ────────────────────

BLOQUES_RESULTADO = [
    {"nombre": "Cartola Personal", "bloque": "cartola",   "col_inicio": 1,  "col_fin": 7},
    {"nombre": "Libro del Banco",  "bloque": "libro",     "col_inicio": 8,  "col_fin": 14},
    {"nombre": "Resultado",        "bloque": "resultado", "col_inicio": 15, "col_fin": 24},
]

BLOQUES_SIN_CONCILIAR = [
    {"nombre": "Cartola Personal", "bloque": "cartola",     "col_inicio": 1, "col_fin": 7},
    {"nombre": "Diagnóstico",      "bloque": "diagnostico", "col_inicio": 8, "col_fin": 10},
]


# ─── Utilidades internas ──────────────────────────────────────────────────────

def _borde_fino() -> Border:
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)


def _borde_medio() -> Border:
    """Borde más grueso para separar bloques visualmente."""
    lado_grueso = Side(style="medium", color="FFFFFF")
    lado_fino   = Side(style="thin",   color="BFBFBF")
    return Border(
        left=lado_grueso,
        right=lado_grueso,
        top=lado_fino,
        bottom=lado_fino,
    )